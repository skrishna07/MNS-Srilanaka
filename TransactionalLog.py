import mysql.connector
from openpyxl import Workbook, load_workbook
import os
import datetime
import shutil
# Function to fetch data from the cPanel database


def fetch_data_from_database(db_config):
    connection = mysql.connector.connect(**db_config)

    cursor = connection.cursor()

    # Example query, replace it with your actual query
    # query = "SELECT OrderID, CINNumber, CompanyName, BotFinalStatus, PaymentByUser, Remarks FROM orders"
    query = "SELECT * from orders WHERE DATE(created_date) = CURDATE() and (LOWER(process_status) = 'completed' or LOWER(process_status) = 'inprogress' or LOWER(process_status) = 'exception')"
    cursor.execute(query)

    data = cursor.fetchall()

    cursor.close()
    connection.close()

    return data

# Function to insert data into Excel file


def generate_transactional_log(db_config,config_dict):
    try:
        config_excel_file_path = config_dict['transactional_log_config_excel_path']
        root_path = config_dict['Root path']
        data = fetch_data_from_database(db_config)
        if not os.path.exists(config_excel_file_path):
            raise Exception(f"Config Excel file not found")
        current_date = datetime.date.today()
        today_date = current_date.strftime("%d-%m-%Y")
        transactional_log_folder = os.path.join(root_path, 'Transactional Log')
        if not os.path.exists(transactional_log_folder):
            os.makedirs(transactional_log_folder)
        transactional_log_file_name = 'TransactionalLog_' + today_date + '.xlsx'
        transactional_log_file_path = os.path.join(transactional_log_folder, transactional_log_file_name)
        shutil.copy(config_excel_file_path, transactional_log_file_path)
        workbook = load_workbook(transactional_log_file_path)

        # Select the active sheet (create a new one if it doesn't exist)
        sheet = workbook.active

        # Add headers if the sheet is new
        if sheet.max_row == 0:
            headers = ['Order ID', 'Registration Number', 'Company Name', 'Bot Final Status', 'Remarks']
            sheet.append(headers)

        # Insert data into the Excel file
        for row in data:
            # Customize the indices based on your actual database columns
            row_data = [row[1], row[3], row[2], row[7], row[11]]
            sheet.append(row_data)

        # Save the workbook
        workbook.save(transactional_log_file_path)
        return transactional_log_file_path
    except Exception as e:
        print(f"Error in generating Transactional Log {e}")
        return None

