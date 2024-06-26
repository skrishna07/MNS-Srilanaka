import openpyxl
import logging
import os
class ConfigException(Exception):
    pass


def create_main_config_dictionary(path, sheet_name):
    try:
        dict_config_main = {}
        if not os.path.exists(path):
            raise ConfigException("Exception: Config file is not exist in the path provided {0}".format(path))
        try:
            logging.info("Config file path: {0} is exist".format(path))
            logging.info("Config file path: {0} is exist".format(path))
            workbook = openpyxl.load_workbook(path)
            logging.info("Config file is loaded")
        except Exception as config_file_error:
            exception_message = ("Exception is occurred while loading "
                                 "Config file from path {0} because {1}").format(path, config_file_error)
            raise ConfigException(exception_message)

        try:
            worksheet = workbook[sheet_name]
            logging.info("Config sheet '{0}' is read".format(sheet_name))
        except Exception as work_sheet_exception:
            exception_message = "Exception is occurred while loading Config Excel file sheet {0} because {1}".format(
                sheet_name, work_sheet_exception)
            raise ConfigException(exception_message)

        maximum_row = worksheet.max_row
        maximum_col = worksheet.max_column

        for config_details in worksheet.iter_rows(min_row=2, min_col=1, max_row=maximum_row, max_col=maximum_col):
            key = config_details[0].value
            value = config_details[1].value
            dict_config_main[key] = value
        logging.info("Config dictionary is created from config file")
        logging.debug(dict_config_main)

        try:
            workbook.close()
            Status="Pass"
        except Exception as config_save_exception:
            exception_message = ("Exception occurred while closing Config file in path {0}"
                                 " because {1}").format(path, config_save_exception)
            raise ConfigException(exception_message)

        return dict_config_main,Status

    except Exception as config_file_read_error:
        logging.critical(config_file_read_error)
        logging.exception(config_file_read_error)
        raise ConfigException(config_file_read_error)